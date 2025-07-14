#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
测试用例生成工具
用于快速创建测试用例模板并添加到Excel文件中
"""

import argparse
import configparser
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class TestCaseGenerator:
    """测试用例生成器类"""
    
    # 默认的测试用例模板
    DEFAULT_TEMPLATE = {
        'test_case_id': '',
        'module': '',
        'name': '',
        'description': '',
        'method': 'GET',
        'path': '',
        'headers': '{"Content-Type": "application/json"}',
        'params': '{}',
        'body': '{}',
        'extract_vars': '{}',
        'asserts': '[{"type": "status_code", "expected": 200}]',
        'pre_condition_tc': '',
        'priority': 'P2',
        'tags': '',
        'is_run': True
    }
    
    # 列宽设置
    COLUMN_WIDTHS = {
        'test_case_id': 15,
        'module': 15,
        'name': 25,
        'description': 40,
        'method': 8,
        'path': 30,
        'headers': 30,
        'params': 30,
        'body': 40,
        'extract_vars': 30,
        'asserts': 40,
        'pre_condition_tc': 15,
        'priority': 8,
        'tags': 15,
        'is_run': 8
    }
    
    def __init__(self, excel_file=None):
        """初始化测试用例生成器
        
        Args:
            excel_file: Excel文件路径，如果不存在则创建
        """
        # 读取配置
        config = configparser.ConfigParser()
        config_file = 'config/config.ini'
        if os.path.exists(config_file):
            config.read(config_file)
            self.excel_file = excel_file or config['TEST'].get('excel_file', 'data/test_cases.xlsx')
        else:
            self.excel_file = excel_file or 'data/test_cases.xlsx'
        
        # 确保目录存在
        os.makedirs(os.path.dirname(self.excel_file), exist_ok=True)
        
        # 加载或创建Excel文件
        if os.path.exists(self.excel_file):
            self.wb = openpyxl.load_workbook(self.excel_file)
            if 'TestCases' not in self.wb.sheetnames:
                self._create_test_case_sheet()
            self.ws = self.wb['TestCases']
        else:
            self.wb = openpyxl.Workbook()
            self.wb.remove(self.wb.active)  # 删除默认创建的sheet
            self._create_test_case_sheet()
    
    def _create_test_case_sheet(self):
        """创建测试用例工作表并设置表头"""
        self.ws = self.wb.create_sheet('TestCases')
        
        # 设置表头
        headers = list(self.DEFAULT_TEMPLATE.keys())
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            self.ws.column_dimensions[get_column_letter(col)].width = self.COLUMN_WIDTHS.get(header, 15)
        
        # 冻结首行
        self.ws.freeze_panes = 'A2'
    
    def _get_next_test_case_id(self, prefix='TC'):
        """获取下一个测试用例ID
        
        Args:
            prefix: 测试用例ID前缀
            
        Returns:
            下一个可用的测试用例ID
        """
        max_id = 0
        for row in range(2, self.ws.max_row + 1):
            test_case_id = self.ws.cell(row=row, column=1).value
            if test_case_id and test_case_id.startswith(prefix):
                try:
                    id_num = int(test_case_id[len(prefix):])
                    max_id = max(max_id, id_num)
                except ValueError:
                    pass
        
        return f"{prefix}{max_id + 1:04d}"
    
    def _get_column_index(self, column_name):
        """获取列名对应的索引
        
        Args:
            column_name: 列名
            
        Returns:
            列索引（从1开始）
        """
        for col in range(1, self.ws.max_column + 1):
            if self.ws.cell(row=1, column=col).value == column_name:
                return col
        return None
    
    def create_test_case(self, test_case_data):
        """创建新的测试用例
        
        Args:
            test_case_data: 测试用例数据字典
            
        Returns:
            创建的测试用例ID
        """
        # 合并默认模板和提供的数据
        test_case = self.DEFAULT_TEMPLATE.copy()
        test_case.update(test_case_data)
        
        # 如果没有提供测试用例ID，则自动生成
        if not test_case['test_case_id']:
            module_prefix = test_case['module'][:2].upper() if test_case['module'] else 'TC'
            test_case['test_case_id'] = self._get_next_test_case_id(prefix=module_prefix)
        
        # 添加到工作表
        next_row = self.ws.max_row + 1
        headers = list(self.DEFAULT_TEMPLATE.keys())
        
        for col, header in enumerate(headers, start=1):
            value = test_case[header]
            self.ws.cell(row=next_row, column=col, value=value)
        
        # 保存工作簿
        self.wb.save(self.excel_file)
        
        return test_case['test_case_id']
    
    def create_test_cases_batch(self, test_cases):
        """批量创建测试用例
        
        Args:
            test_cases: 测试用例数据字典列表
            
        Returns:
            创建的测试用例ID列表
        """
        test_case_ids = []
        
        for test_case_data in test_cases:
            test_case_id = self.create_test_case(test_case_data)
            test_case_ids.append(test_case_id)
        
        return test_case_ids
    
    def create_test_suite(self, suite_name, test_cases):
        """创建测试套件（一组相关的测试用例）
        
        Args:
            suite_name: 测试套件名称（将用作模块名）
            test_cases: 测试用例数据字典列表，每个字典至少包含name和path
            
        Returns:
            创建的测试用例ID列表
        """
        enriched_test_cases = []
        
        for i, tc in enumerate(test_cases):
            # 确保每个测试用例都有模块名
            tc['module'] = tc.get('module', suite_name)
            
            # 如果没有提供测试用例ID，则自动生成
            if not tc.get('test_case_id'):
                module_prefix = tc['module'][:2].upper()
                tc['test_case_id'] = f"{module_prefix}{i+1:04d}"
            
            enriched_test_cases.append(tc)
        
        return self.create_test_cases_batch(enriched_test_cases)
    
    def generate_crud_suite(self, resource_name, base_path):
        """生成CRUD（创建、读取、更新、删除）测试套件
        
        Args:
            resource_name: 资源名称（如users, products等）
            base_path: 基础路径（如/api/v1/）
            
        Returns:
            创建的测试用例ID列表
        """
        # 确保base_path以/开头且以/结尾
        if not base_path.startswith('/'):
            base_path = '/' + base_path
        if not base_path.endswith('/'):
            base_path = base_path + '/'
        
        # 构建资源路径
        resource_path = f"{base_path}{resource_name}"
        
        # 创建CRUD测试用例
        test_cases = [
            # 创建资源
            {
                'name': f"创建{resource_name}",
                'description': f"测试创建新的{resource_name}资源",
                'method': 'POST',
                'path': resource_path,
                'body': '{"name": "测试名称", "description": "测试描述"}',
                'extract_vars': '{"resource_id": "$.id"}',
                'asserts': '[{"type": "status_code", "expected": 201}, {"type": "jsonpath", "expression": "$.id", "expected_not": null}]',
                'priority': 'P1',
                'tags': 'create,smoke'
            },
            # 获取资源列表
            {
                'name': f"获取{resource_name}列表",
                'description': f"测试获取{resource_name}资源列表",
                'method': 'GET',
                'path': resource_path,
                'params': '{"limit": 10, "offset": 0}',
                'asserts': '[{"type": "status_code", "expected": 200}, {"type": "jsonpath", "expression": "$", "expected_type": "array"}]',
                'priority': 'P1',
                'tags': 'read,smoke'
            },
            # 获取单个资源
            {
                'name': f"获取单个{resource_name}",
                'description': f"测试获取单个{resource_name}资源",
                'method': 'GET',
                'path': f"{resource_path}/{{resource_id}}",
                'pre_condition_tc': f"{resource_name.upper()[:2]}0001",  # 依赖创建资源的测试用例
                'asserts': '[{"type": "status_code", "expected": 200}, {"type": "jsonpath", "expression": "$.id", "expected": "{{resource_id}}"}]',
                'priority': 'P1',
                'tags': 'read,smoke'
            },
            # 更新资源
            {
                'name': f"更新{resource_name}",
                'description': f"测试更新{resource_name}资源",
                'method': 'PUT',
                'path': f"{resource_path}/{{resource_id}}",
                'body': '{"name": "更新的名称", "description": "更新的描述"}',
                'pre_condition_tc': f"{resource_name.upper()[:2]}0001",  # 依赖创建资源的测试用例
                'asserts': '[{"type": "status_code", "expected": 200}, {"type": "jsonpath", "expression": "$.name", "expected": "更新的名称"}]',
                'priority': 'P2',
                'tags': 'update'
            },
            # 部分更新资源
            {
                'name': f"部分更新{resource_name}",
                'description': f"测试部分更新{resource_name}资源",
                'method': 'PATCH',
                'path': f"{resource_path}/{{resource_id}}",
                'body': '{"description": "部分更新的描述"}',
                'pre_condition_tc': f"{resource_name.upper()[:2]}0001",  # 依赖创建资源的测试用例
                'asserts': '[{"type": "status_code", "expected": 200}, {"type": "jsonpath", "expression": "$.description", "expected": "部分更新的描述"}]',
                'priority': 'P2',
                'tags': 'update'
            },
            # 删除资源
            {
                'name': f"删除{resource_name}",
                'description': f"测试删除{resource_name}资源",
                'method': 'DELETE',
                'path': f"{resource_path}/{{resource_id}}",
                'pre_condition_tc': f"{resource_name.upper()[:2]}0001",  # 依赖创建资源的测试用例
                'asserts': '[{"type": "status_code", "expected": 204}]',
                'priority': 'P1',
                'tags': 'delete,smoke'
            },
            # 验证资源已删除
            {
                'name': f"验证{resource_name}已删除",
                'description': f"测试验证{resource_name}资源已被删除",
                'method': 'GET',
                'path': f"{resource_path}/{{resource_id}}",
                'pre_condition_tc': f"{resource_name.upper()[:2]}0006",  # 依赖删除资源的测试用例
                'asserts': '[{"type": "status_code", "expected": 404}]',
                'priority': 'P2',
                'tags': 'delete'
            }
        ]
        
        return self.create_test_suite(resource_name, test_cases)


def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description="测试用例生成工具")
    
    # 子命令
    subparsers = parser.add_subparsers(dest="command", help="子命令")
    
    # 创建单个测试用例
    create_parser = subparsers.add_parser("create", help="创建单个测试用例")
    create_parser.add_argument("-n", "--name", required=True, help="测试用例名称")
    create_parser.add_argument("-m", "--module", required=True, help="模块名称")
    create_parser.add_argument("-d", "--description", help="测试用例描述")
    create_parser.add_argument("--method", default="GET", choices=["GET", "POST", "PUT", "DELETE", "PATCH"], help="HTTP方法")
    create_parser.add_argument("-p", "--path", required=True, help="API路径")
    create_parser.add_argument("--body", help="请求体JSON字符串")
    create_parser.add_argument("--priority", default="P2", choices=["P0", "P1", "P2", "P3", "P4"], help="优先级")
    create_parser.add_argument("--tags", help="标签，用逗号分隔")
    
    # 生成CRUD测试套件
    crud_parser = subparsers.add_parser("crud", help="生成CRUD测试套件")
    crud_parser.add_argument("-r", "--resource", required=True, help="资源名称")
    crud_parser.add_argument("-b", "--base-path", default="/api/v1/", help="基础路径")
    
    # 通用参数
    parser.add_argument("-f", "--file", help="Excel文件路径")
    
    return parser.parse_args()


def main():
    """主函数"""
    args = parse_arguments()
    
    try:
        generator = TestCaseGenerator(args.file)
        
        if args.command == "create":
            # 创建单个测试用例
            test_case_data = {
                'name': args.name,
                'module': args.module,
                'description': args.description or args.name,
                'method': args.method,
                'path': args.path,
                'priority': args.priority,
                'tags': args.tags or ''
            }
            
            if args.body:
                test_case_data['body'] = args.body
            
            test_case_id = generator.create_test_case(test_case_data)
            print(f"成功创建测试用例: {test_case_id}")
            
        elif args.command == "crud":
            # 生成CRUD测试套件
            test_case_ids = generator.generate_crud_suite(args.resource, args.base_path)
            print(f"成功创建CRUD测试套件，包含 {len(test_case_ids)} 个测试用例:")
            for tc_id in test_case_ids:
                print(f"  - {tc_id}")
        
        else:
            print("请指定要执行的命令。使用 -h 查看帮助信息。")
            return 1
        
        print(f"\n测试用例已保存到: {os.path.abspath(generator.excel_file)}")
        return 0
        
    except Exception as e:
        print(f"错误: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())